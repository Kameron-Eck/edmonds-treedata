import sys, json, openpyxl
from openpyxl.styles import Alignment
PATH = r'G:/My Drive/treedata/Literature_Tracker.xlsx'
rows = json.load(open(sys.argv[1], encoding='utf-8'))
phase = json.load(open(sys.argv[2], encoding='utf-8')) if len(sys.argv) > 2 else []
wb = openpyxl.load_workbook(PATH)
ws = wb['Literature Tracker']
# next id
ids = [c.value for c in ws['A'][1:] if isinstance(c.value, int)]
nid = max(ids) + 1
ref = ws['A2']  # style template
for r in rows:
    out = [nid, r['authors'], int(r['year']), r['title'], r['source'], r['relevance'], r['phase'], r['url'], 'To Read']
    ws.append(out)
    for col in range(1, 10):
        c = ws.cell(row=ws.max_row, column=col)
        c.alignment = Alignment(wrap_text=True, vertical='top')
    print(nid, '|', r['authors'], r['year'], '|', r['title'][:70])
    nid += 1
if phase:
    ws2 = wb['Search Phase Reference']
    for p in phase:
        ws2.append([p['phase'], p['topic'], 'Complete'])
        print('PHASE:', p['phase'], p['topic'])
wb.save(PATH)
print('SAVED. total rows:', ws.max_row)
