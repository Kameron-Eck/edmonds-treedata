"""Record a campaign acquisition in imagery_catalog_2026-08-22.xlsx WITHOUT rebuilding it (cell edits in the
sheet-adder style; a build_master_catalog.py rebuild would drop the ten appended sheets).

  MASTER:               the row whose 'Data-lake file' == --old gets 'Replace with a better version?' stamped;
                        a NEW row for --new is appended (Already in data lake? = YES, Evidence = MEASURED).
  Effective_Resolution: one row for --new (from the verify measure.json).
  DataLake_Issues:      the --old row's 'Replace with better?' cell stamped RESOLVED.
Idempotent on the new file name.

  PYTHONUTF8=1 py -3.12 scratch/imagery_catalog_flip.py --id S16 --old 2016_snoh_rgbi.tif --new 2016_snoh_1ft_rgbi.tif
"""
import argparse, datetime as dt, json, pathlib, sys
from openpyxl import load_workbook

SCRIPTS = pathlib.Path(__file__).resolve().parent.parent
XLSX = SCRIPTS / "imagery_catalog_2026-08-22.xlsx"
LOCAL = pathlib.Path(r"D:\edmonds-pipeline\Imagery")
TODAY = dt.date.today().isoformat()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--id", required=True); ap.add_argument("--old"); ap.add_argument("--new", required=True)
    ap.add_argument("--source-dir", default="SnoCo"); ap.add_argument("--product", default="")
    a = ap.parse_args([x for x in sys.argv[1:] if not (x == "-f" or x.endswith(".json"))])
    acq = LOCAL / a.source_dir / "_acq" / a.id
    meas = json.loads((acq / "measure.json").read_text(encoding="utf-8"))
    dec = json.loads((acq / "decision.json").read_text(encoding="utf-8"))
    wb = load_workbook(XLSX)

    # --- MASTER
    ws = wb["MASTER"]; hdr = [c.value for c in ws[1]]; col = {h: i + 1 for i, h in enumerate(hdr)}
    present = any(str(r[col["Data-lake file"] - 1].value) == a.new for r in ws.iter_rows(min_row=2))
    if a.old:
        for r in ws.iter_rows(min_row=2):
            if str(r[col["Data-lake file"] - 1].value) == a.old:
                cell = r[col["Replace with a better version?"] - 1]
                if a.new not in str(cell.value or ""):
                    cell.value = (f"{dec['verdict']} {TODAY} -> {a.new} ({'; '.join(dec.get('wins', []))}). " + str(cell.value or ""))[:2000]
    if not present:
        row = {h: "" for h in hdr}
        row.update({"Category": "imagery", "Year": meas.get("file", "")[:4], "Product name": a.product or f"{a.id} campaign acquisition ({a.new})",
                    "Source": "acquire_imagery campaign 2026-08-23", "Acquisition date": "see Pixel_Size_And_Date row",
                    "Metadata (key facts)": f"true GSD {meas['true_gsd_cm']} cm; {meas['bands']} bands; effective {meas.get('effective_cm')} cm; EPSG:{meas['epsg']}; city coverage {meas.get('city_coverage_pct')}%",
                    "Link to metadata": str(meas.get("tags", {}).get("SOURCE_URL", "")), "Link to download": str(meas.get("tags", {}).get("SOURCE_URL", "")),
                    "Already in data lake?": "YES", "Data-lake file": a.new, "Issue with the data-lake copy?": "none known (measured on arrival)",
                    "Complements the data lake?": "n/a (already held)", "Replace with a better version?": f"this file {dec['verdict']}S {a.old}" if a.old else "no",
                    "Evidence": "MEASURED", "Notes": f"decision {dec['verdict']}: wins={dec.get('wins')} losses={dec.get('losses')}; measure.json under {acq}"})
        ws.append([row[h] for h in hdr])

    # --- Effective_Resolution
    ws = wb["Effective_Resolution"]
    if not any(str(r[0].value) == a.new for r in ws.iter_rows(min_row=2)):
        ws.append([a.new, a.source_dir, meas["true_gsd_cm"], meas.get("oversampling"), meas.get("effective_cm"), f"NEW {TODAY} (acquire_imagery verify)",
                   f"median of {meas.get('n_sites')} Method_Provenance sites; common-grid vs {a.old}: new {((meas.get('compare_to_held') or {}).get('effective_cm_common_new'))} / held {((meas.get('compare_to_held') or {}).get('effective_cm_common_held'))} cm, HF ratio {((meas.get('compare_to_held') or {}).get('hf_ratio_new_over_held'))}" if a.old else f"median of {meas.get('n_sites')} sites"])

    # --- DataLake_Issues
    if a.old and "DataLake_Issues" in wb.sheetnames:
        ws = wb["DataLake_Issues"]
        for r in ws.iter_rows(min_row=2):
            if str(r[0].value) == a.old and a.new not in str(r[2].value or ""):
                r[2].value = f"RESOLVED {TODAY}: {dec['verdict']} by {a.new}. " + str(r[2].value or "")
    wb.save(XLSX)
    print(f"catalogue updated: MASTER {'(row existed)' if present else '+1 row'}, Effective_Resolution, DataLake_Issues for {a.new}")


if __name__ == "__main__":
    main()
