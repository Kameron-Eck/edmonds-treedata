"""Add/replace the Pixel_Size_And_Date sheet in imagery_catalog_2026-08-22.xlsx from qc/imagery_pixelsize_and_date.csv.

Deliberately NOT a full rebuild: build_master_catalog.py writes a fresh Workbook() containing only its own sheets, and the
ten sheets appended afterwards (Processing_Mods, Pixel_Forensics, Effective_Resolution, Provenance_Chain, Dating_Methods,
Reference_Independence, Retractions, Method_Provenance, Contradictions, Licensing_Risk) would be destroyed by re-running it.
This follows the add_*_sheets.py pattern instead (delete-if-exists, then append), with the same citation gate applied in
imagery_pixelsize_date_build.py. Run after the CSV builder:
    PYTHONUTF8=1 py -3.12 scratch/imagery_pixelsize_date_build.py && PYTHONUTF8=1 py -3.12 scratch/imagery_pixelsize_date_sheet.py
"""
import csv, pathlib
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = pathlib.Path(__file__).resolve().parent
XLSX = HERE.parent / "imagery_catalog_2026-08-22.xlsx"
CSV = HERE.parent / "qc" / "imagery_pixelsize_and_date.csv"
SHEET = "Pixel_Size_And_Date"
HF, HFONT = PatternFill("solid", fgColor="1F4E79"), Font(bold=True, color="FFFFFF", size=10)
GRADE_FILL = {"MEASURED": "C6EFCE", "PUBLISHED": "DDEBF7", "INFERRED": "FFEB9C", "NOT FOUND": "F8CBAD"}

rows = list(csv.DictReader(open(CSV, encoding="utf-8")))
headers = list(rows[0].keys())
wb = load_workbook(XLSX)
if SHEET in wb.sheetnames:
    del wb[SHEET]
ws = wb.create_sheet(SHEET, index=1)   # right after README
ws.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c); cell.fill, cell.font = HF, HFONT
    cell.alignment = Alignment(vertical="center", wrap_text=True)
for r in rows:
    ws.append([(r.get(h) or "")[:2000] for h in headers])
widths = {"file": 28, "year_label": 12, "source": 40, "grid_px": 10, "grid_units": 18, "crs": 12, "true_ground_cm": 10,
          "effective_cm": 18, "native_flight_cm": 34, "date_shot": 44, "date_precision": 18, "single_or_multi_date": 26,
          "evidence_grade": 11, "source_url": 44, "verbatim_quote": 50, "notes": 70, "px_evidence": 30,
          "alt_date_shot": 30, "alt_source_url": 36, "alt_verbatim_quote": 40, "row_type": 14}
for i, h in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 20)
ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
gi = headers.index("evidence_grade") + 1
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    g = str(row[gi - 1].value or "")
    if g in GRADE_FILL:
        row[gi - 1].fill = PatternFill("solid", fgColor=GRADE_FILL[g])
# README pointer (one line, idempotent)
rd = wb["README"]
marker = "Pixel_Size_And_Date:"
if not any(marker in str(c.value or "") for r in rd.iter_rows() for c in r):
    rd.append([f"{marker} true pixel size (grid / true ground / effective / native flight) + date shot per held acquisition, "
               f"each with evidence grade, fetched URL and verbatim quote. Source of record = Scripts/qc/imagery_pixelsize_and_date.csv "
               f"(built by scratch/imagery_pixelsize_date_build.py; this sheet is a copy)."])
wb.save(XLSX)
print(f"{SHEET}: {len(rows)} rows written to {XLSX.name}; sheets now {len(wb.sheetnames)}")
