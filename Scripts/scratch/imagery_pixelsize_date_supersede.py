"""Mark catalogue rows superseded by the 2026-08-23 Pixel_Size_And_Date work (in place; no rebuild).

Touches only: Contradictions (rows '2015 flight date', '2017 imagery source'), Provenance_Chain (2017/2020/2022 rows
and the dead /arcgis/ Marsh path), Retractions (two new rows), Dating_Methods (method 8 note). Idempotent.
"""
import pathlib
from openpyxl import load_workbook
XLSX = pathlib.Path(__file__).resolve().parent.parent / "imagery_catalog_2026-08-22.xlsx"
TAG = "[SUPERSEDED 2026-08-23 -> Pixel_Size_And_Date]"
wb = load_workbook(XLSX)

def stamp(ws, key_col, key_startswith, col_name, text):
    hdr = [c.value for c in ws[1]]
    ki, ci = hdr.index(key_col) + 1, hdr.index(col_name) + 1
    for row in ws.iter_rows(min_row=2):
        v = str(row[ki - 1].value or "")
        if v.startswith(key_startswith):
            cur = str(row[ci - 1].value or "")
            if TAG not in cur:
                row[ci - 1].value = (TAG + " " + text + " || was: " + cur)[:2000]
            return True
    return False

ws = wb["Contradictions"]
stamp(ws, "Subject", "2015 flight date", "Status",
      "LARGELY RESOLVED: the three dates are three DIFFERENT products. Aug 7 2015 = NAIP/HXIP 2015 (Davey UTC Assessment names USDA FSA), "
      "Feb 15-Mar 8 = King Pictometry (the held 2015_king_rgb.tif; pixel-distinct from the CoE basemap, r 0.04-0.56 vs 0.99 control), "
      "Apr 8 + Apr 17 2015 = the city consortium photo centres (no Apr 9 over Edmonds). Residual: attribution of Ed_Aerial_2015.tif to the consortium is INFERRED; the '4.8 in' figure is unexplained.")
stamp(ws, "Subject", "2017 imagery source", "Status",
      "RESOLVED: 2017_coe_rgb.tif is the SAME ORTHOMOSAIC as King County IMAGE_Ortho2017KCNAT (Pictometry), MEASURED r 0.957-0.997 over 66 citywide windows; dated 2017-05-04..05-10 from the per-frame index. "
      "The council-minutes 'PlanIt Geo' line is retired (the 2019 UFMP was by Davey Resource Group and used Aug-2015 imagery).")

ws = wb["Provenance_Chain"]
stamp(ws, "Target", "2017_coe_rgb.tif", "Status", "DATED: same orthomosaic as King 2017 -> 2017-05-04..05-10 MEASURED.")
stamp(ws, "Target", "2022_coe_rgb.tif", "Status", "IDENTITY MEASURED: same MrSID encode as Everett's Image2022_3in_sid (input set 675,609,635,568 B) -> window PUBLISHED.")
stamp(ws, "Target", "2020_coe_rgb.tif", "Status", "Still INFERRED; consistent with ONE continuous pass (shadow geometry, 10 sites); head trimmed to 2020-04-25. Contract vehicle for the PRR: King County RFP 1166-18-PCR / Piggyback PB-19-14BC.")
stamp(ws, "Target", "Edmonds_Marsh_2018", "Document / endpoint",
      "LIVE PATH: https://maps.edmondswa.gov/gis/rest/services/Basemap/Edmonds_Marsh_2018/ImageServer/keyProperties?f=json (the /arcgis/ form 500s; /info/keyProperties 400s on every service).")

ws = wb["Dating_Methods"]
stamp(ws, "#", "8", "Why it matters / caveat", "Contract numbers for the ask/PRR now known: King County RFP 1166-18-PCR, Snohomish Piggyback PB-19-14BC, Legistar 2024-1168.")
stamp(ws, "#", "1", "Why it matters / caveat", "RUN 2026-08-23: azimuth reliable (2012 control 8/10), elevation saturates above ~45 deg -> 2020 consistent with one pass, no calendar date. Evidence qc/imagery_date_evidence/shadow_dating_2020/.")
stamp(ws, "#", "2", "Why it matters / caveat", "RUN 2026-08-23: the 30-deg sun floor is FALSIFIED by a known flight day (2015-02-15, max elevation 29.3 deg); calibrated rule (20-deg floor, 3-station consensus) passes 46/46 controls. Eliminates 22/92 (2020), 19/97 (2022), 18/62 (2024), 32/140 (2021s) days.")

ws = wb["Retractions"]
hdr = [c.value for c in ws[1]]
existing = {str(r[0].value or "") for r in ws.iter_rows(min_row=2)}
NEW = [
 ["2017_coe_rgb.tif spans an 8.5-month leaf-off-to-leaf-on window (2017-02-17 to 10-30)", "2026-08-23", "IMAGERY_FACTS 9.2 / Provenance_Chain",
  "It is the same orthomosaic as 2017_king_rgb.tif (r 0.957-0.997, 66 windows) and was flown 2017-05-04..05-10 over the city.",
  "A coarse published window for a PARENT product was carried as the file's own window. Measuring identity against a sibling file settled in minutes what document-hunting could not."],
 ["The 2015 CoE basemap date is a three-way contest between independent acquisitions", "2026-08-23", "Contradictions sheet",
  "The three dates belong to three DIFFERENT products (NAIP/HXIP Aug 7; King Pictometry Feb-Mar; city consortium Apr 8/17). The catalogue's 'Apr 8/9/17' also contained a day (Apr 9) with no frames over Edmonds.",
  "Year-matched product substitution is the recurring trap (it recurred the same day: a second 2020 acquisition, Hexagon 2020-08-27/28, exists over Edmonds). Always ask WHICH product before asking WHEN."],
 ["The HXIP footprint SDATE '08/11/2016 16:52' is UTC because the held file shows morning shadows", "2026-08-23", "session lead's open note",
  "SDATE is LOCAL time (7034-footprint SUNEL fit; 09:05 read as UTC puts the sun 26 deg below the horizon). The held county 1-ft product simply is not that sortie.",
  "A hypothesis stated to explain one observation was tested against the whole field and failed. State it, test it, record it."],
]
for r in NEW:
    if r[0] not in existing:
        ws.append(r)
wb.save(XLSX)
print("supersessions stamped; sheets:", len(wb.sheetnames))
